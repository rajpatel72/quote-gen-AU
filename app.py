"""
Streamlit app: Electricity bill -> fill fixed Excel template (Current Energy Offer)
Columns used (fixed):
 - Units  -> column A
 - Description -> column C
 - Before Discount (rate) -> column D
 - Conditional Discount -> column E

This script:
 - accepts bill (pdf/image) + template xlsx (optional)
 - extracts text (pdfplumber, fallback to pytesseract)
 - parses header fields and dynamic usage lines
 - displays parsed lines for editing
 - writes results into the provided template and returns filled xlsx for download
"""

import io
import re
from typing import List, Dict, Tuple


import streamlit as st
import pdfplumber
from PIL import Image, ImageOps, ImageFilter
import pytesseract
from openpyxl import load_workbook
import pandas as pd

# Optional: if you want better PDF->image fallback, install pdf2image and poppler on the host.
try:
    from pdf2image import convert_from_bytes
    PDF2IMAGE_AVAILABLE = True
except Exception:
    PDF2IMAGE_AVAILABLE = False

# -----------------------
# Utilities
# -----------------------
def ocr_image(pil_image: Image.Image) -> str:
    """Basic preprocessing + pytesseract OCR"""
    # Convert to greyscale, enhance
    img = pil_image.convert("L")
    # Resize up if small (helps OCR)
    max_dim = 2000
    if max(img.size) < 1500:
        scale = int(max_dim / max(img.size))
        if scale > 1:
            img = img.resize((img.width * scale, img.height * scale), Image.LANCZOS)
    # Slight sharpening
    img = img.filter(ImageFilter.SHARPEN)

    # Optional binarization (works sometimes): uncomment if needed
    # img = img.point(lambda x: 0 if x < 160 else 255, '1')

    # OCR config; page segmentation mode 6 tends to be good for blocks of text.
    cfg = "--psm 6"
    text = pytesseract.image_to_string(img, config=cfg)
    return text


def extract_text_from_pdf_bytes(file_bytes: bytes) -> str:
    """Try to extract text from a PDF. Use pdfplumber for text; if very little text, fallback to OCR per-page."""
    text = ""
    try:
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            for page in pdf.pages:
                page_text = (page.extract_text() or "").strip()
                if page_text:
                    text += page_text + "\n"
    except Exception as e:
        st.warning(f"pdfplumber error: {e}")

    # if we got good text, return it
    if len(text.strip()) > 100:
        return text

    # fallback: convert pages to images + OCR (requires pdf2image/poppler)
    if PDF2IMAGE_AVAILABLE:
        try:
            images = convert_from_bytes(file_bytes, dpi=300)
            page_texts = []
            for im in images:
                page_texts.append(ocr_image(im))
            return "\n".join(page_texts)
        except Exception as e:
            st.warning(f"pdf2image fallback failed: {e}")
            # fallback to returning whatever text we had (maybe empty)
    else:
        st.info("pdf2image/poppler not available — falling back to pdfplumber only.")

    return text


def extract_text_from_image_bytes(file_bytes: bytes) -> str:
    """Open bytes as image and OCR."""
    try:
        im = Image.open(io.BytesIO(file_bytes))
    except Exception as e:
        st.error(f"Cannot open image: {e}")
        return ""
    return ocr_image(im)


# ---------------
# Parsing logic
# ---------------
def find_first_regex(text: str, patterns) -> str:
    """Try multiple regex patterns and return first group's content or empty."""
    for pat in patterns:
        m = re.search(pat, text, re.IGNORECASE | re.MULTILINE)
        if m:
            # return the first capturing group if present else full match
            if m.lastindex:
                return m.group(1).strip()
            return m.group(0).strip()
    return ""


def parse_header_fields(text: str) -> Dict[str, str]:
    """Extract a set of header fields from bill text using multiple heuristics."""
    out = {}
    # NMI (common formats: digits maybe 11-12 digits)
    out["NMI"] = find_first_regex(text, [r"NMI[:\s]*([A-Z0-9\-]+)", r"NMI/MIRN[:\s]*([0-9\-]+)", r"\b(\d{10,13})\b"])
    # Account Number
    out["Account Number"] = find_first_regex(text, [r"Account\s*Number[:\s]*([A-Z0-9\-]+)", r"Account\s*No[:\s]*([A-Z0-9\-]+)"])
    # Customer name / Account name
    out["Customer Name"] = find_first_regex(text, [r"Account\s*Name[:\s]*([A-Z\-\&\.\,\s0-9]+)", r"Customer[:\s]*([A-Z\-\&\.\,\s0-9]+)", r"Bill to[:\s]*([A-Z\-\&\.\,\s0-9]+)"])
    # Retailer
    out["Retailer"] = find_first_regex(text, [r"Retailer[:\s]*([A-Za-z0-9 \-&]+)", r"Current\s*Energy\s*Retailer[:\s]*([A-Za-z0-9 \-&]+)"])
    # Site address (attempt to find lines containing road/st, rd, drive, vic/nsw etc.)
    out["Site Address"] = find_first_regex(text, [r"Service\s*Address[:\s]*([A-Z0-9a-z\,\.\-\/\s]+)", r"Site Address[:\s]*([A-Z0-9a-z\,\.\-\/\s]+)", r"(\d+\s+[A-Za-z0-9\s]+\s+(Road|Rd|Street|St|Drive|Dr|Lane|Ln|Way|Ave|Avenue)[^\n\r]*)"])
    # Billing period
    out["Billing Period"] = find_first_regex(text, [r"Bill\s*Period[:\s]*([\d]{1,2}\s+[A-Za-z]{3,9}\s+\d{4}\s*-\s*[\d]{1,2}\s+[A-Za-z]{3,9}\s+\d{4})", r"Period[:\s]*([\d\/\-\s]+to[\d\/\-\s]+)"])
    # Total amount
    out["Total Charges"] = find_first_regex(text, [r"Total\s*Amount[:\s]*\$?([\d,]+\.\d{2})", r"Total\s*\(GST.*\)[:\s]*\$?([\d,]+\.\d{2})", r"Amount\s*Due[:\s]*\$?([\d,]+\.\d{2})"])
    return out


# Parsing usage/tariff lines:
UNIT_RE = r"(\d{1,3}(?:[,\d]{0,3})?(?:\.\d+)?)"  # matches numbers with commas/decimals
MONEY_RE = r"(\$?\d+\.\d{2,4}|\d+\.\d{2,4}|\d+\.\d{1,4})"  # rate or money
PCT_RE = r"(\d{1,2}(?:\.\d+)?\s*\%)"

COMMON_KEYWORDS = r"\b(peak|off[-\s]?peak|standing|supply|demand|solar|feed|concession|usage|kwh|fixed|daily|annual)\b"


def extract_numbers_from_string(s: str) -> List[str]:
    # return all numbers that look like units or money (strip $ and commas)
    found = re.findall(r"\$?[\d,]+(?:\.\d+)?", s)
    return [x.replace("$", "").replace(",", "") for x in found]


def parse_usage_lines(text: str) -> List[Dict]:
    """
    Heuristic parser to find usage / tariff lines.
    Returns list of dicts: {"Units":..., "Description":..., "Rate":..., "Discount":...}
    """
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    parsed = []

    # Approach:
    # 1) First pass: find lines that contain a keyword and at least one numeric value.
    for ln in lines:
        if re.search(COMMON_KEYWORDS, ln, re.IGNORECASE) and re.search(r"\d", ln):
            # avoid lines that are summary totals (look for 'total' text)
            if re.search(r"\btotal\b", ln, re.IGNORECASE):
                continue

            nums = extract_numbers_from_string(ln)
            discount = ""
            discount_m = re.search(PCT_RE, ln)
            if discount_m:
                discount = discount_m.group(1).replace(" ", "")

            # pick units: choose the largest numeric (in digits) or the leftmost numeric > 0
            units = ""
            rate = ""
            if nums:
                # heuristics: rates often are small like 0.0632 or $0.0632 or numbers with 4 decimals
                possible_rates = [n for n in nums if re.match(r"^\d+\.\d{3,4}$", n) or float(n) < 10]
                if possible_rates:
                    # take the FIRST small number as rate
                    rate = possible_rates[0]
                    # take the largest other number as units (if present)
                    others = [n for n in nums if n != rate]
                    if others:
                        # choose the largest by numeric value as units
                        units = max(others, key=lambda x: float(x))
                    else:
                        units = ""
                else:
                    # fallback: if only two numbers, assume first is units, second is total or rate
                    if len(nums) >= 2:
                        units = nums[0]
                        rate = nums[1]
                    else:
                        units = nums[0]
                        rate = ""
            # description: remove numbers & currency and pct; whatever remains is description
            desc = re.sub(r"\$?[\d,]+(?:\.\d+)?", "", ln)
            desc = re.sub(r"\d+\s*%", "", desc)
            desc = desc.strip(" -:|")

            parsed.append({
                "Units": units,
                "Description": desc,
                "Rate": rate,
                "Discount": discount
            })

    # 2) If no lines found using keywords, try table-ish lines: lines containing >=2 numeric tokens
    if not parsed:
        for ln in lines:
            nums = extract_numbers_from_string(ln)
            if len(nums) >= 2 and not re.search(r"\b(total|gst|balance|amount due)\b", ln, re.IGNORECASE):
                discount = ""
                discount_m = re.search(PCT_RE, ln)
                if discount_m:
                    discount = discount_m.group(1)
                # numeric heuristics
                units = nums[0]
                # try find a small number as rate
                possible_rates = [n for n in nums[1:] if float(n) < 10]
                rate = possible_rates[0] if possible_rates else (nums[1] if len(nums) > 1 else "")
                desc = re.sub(r"\$?[\d,]+(?:\.\d+)?", "", ln).strip(" -:|")
                parsed.append({
                    "Units": units, "Description": desc, "Rate": rate, "Discount": discount
                })

    # dedupe/clean short items
    cleaned = []
    for item in parsed:
        # normalize numbers
        def clean_num(s):
            if not s:
                return ""
            try:
                return str(float(s))
            except Exception:
                # remove commas and $ etc
                s2 = s.replace(",", "").replace("$", "")
                return s2

        item["Units"] = clean_num(item.get("Units", "") or "")
        item["Rate"] = clean_num(item.get("Rate", "") or "")
        item["Discount"] = (item.get("Discount") or "").replace(" ", "")
        item["Description"] = re.sub(r"\s{2,}", " ", item.get("Description", "").strip(",:-"))
        # skip items that are too short or obviously totals
        if item["Description"] and not re.search(r"\b(total|gst|amount due|balance)\b", item["Description"], re.IGNORECASE):
            cleaned.append(item)

    # limit to reasonable number (e.g., 20 lines)
    return cleaned[:30]


# -----------------------
# Excel writing
# -----------------------
def find_header_start(ws) -> Tuple[int, Dict[str, int]]:
    """
    Search workbook (first sheet) for header labels "Units" and "Description".
    Returns header_row (int) and mapping of expected columns to their column index (1-based).
    If not found, returns (None,{}) so caller can fallback to fixed row.
    """
    header_texts = {"Units": None, "Description": None, "Before Discount": None}
    for row in ws.iter_rows(min_row=1, max_col=30, max_row=200):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                v = cell.value.strip().lower()
                if "units" in v and header_texts["Units"] is None:
                    header_texts["Units"] = (cell.row, cell.column)
                if "description" in v and header_texts["Description"] is None:
                    header_texts["Description"] = (cell.row, cell.column)
                if "before" in v and "discount" in v and header_texts["Before Discount"] is None:
                    header_texts["Before Discount"] = (cell.row, cell.column)
    # decide header_row (use the min row among found header cells)
    rows_found = [r for (r, c) in header_texts.values() if r is not None] if any(header_texts.values()) else []
    if rows_found:
        header_row = min(rows_found)
        # build col map: map 'Units'->col index, 'Description'->col index, 'Before'->col index, 'Discount'->col index
        col_map = {}
        # find exact columns by re-searching that header_row
        for cell in ws[header_row]:
            if cell.value and isinstance(cell.value, str):
                v = cell.value.strip().lower()
                if "units" in v:
                    col_map["Units"] = cell.column
                if "description" in v:
                    col_map["Description"] = cell.column
                if "before" in v and "discount" in v:
                    col_map["Before Discount"] = cell.column
                if "conditional" in v and "discount" in v:
                    col_map["Conditional Discount"] = cell.column
        return header_row, col_map
    return None, {}


import io
from openpyxl import load_workbook

def write_usage_to_template(template_bytes, headers, usage_lines_final):
    """
    Fill the Excel template with header details and dynamic usage lines.
    
    Arguments:
      template_bytes: bytes of the Excel template
      headers: dict with fixed fields (Customer Name, NMI, Retailer, etc.)
      usage_lines_final: list of dicts with dynamic usage rows, e.g.
          [
              {"Units": "22491.80", "Description": "Peak", "Rate": "0.0632", "Discount": ""},
              {"Units": "25764.91", "Description": "Off-Peak", "Rate": "0.0355", "Discount": ""},
          ]
    """

    def safe_val(v):
        """Convert values into Excel-safe format"""
        if v is None:
            return ""
        if isinstance(v, (list, dict)):
            return str(v)  # flatten to text
        try:
            # keep numbers as numbers
            if isinstance(v, (int, float)):
                return v
            s = str(v).strip()
            return float(s) if s.replace(".", "", 1).isdigit() else s
        except Exception:
            return str(v)

    # Load workbook from bytes
    input_stream = io.BytesIO(template_bytes)
    wb = load_workbook(input_stream)
    ws = wb.active

    # ---- Header section ----
    ws["A6"]  = safe_val(headers.get("Customer Name"))
    ws["B15"] = safe_val(headers.get("Meter Type"))
    ws["B16"] = safe_val(headers.get("Tariff Classification"))
    ws["B17"] = safe_val(headers.get("Distribution Region"))
    ws["B18"] = safe_val(headers.get("Site Address"))
    ws["B19"] = safe_val(headers.get("NMI"))
    ws["B20"] = safe_val(headers.get("Retailer"))

    # ---- Current Energy Offer table ----
    start_row = 34  # adjust if your template rows shift

    # fixed column indexes
    unit_col   = 1  # A → Units
    desc_col   = 3  # C → Description
    before_col = 4  # D → Before Discount
    disc_col   = 5  # E → Conditional Discount

    # Fill dynamic rows
    for i, line in enumerate(usage_lines_final):
        r = start_row + i

        if not isinstance(line, dict):
            raise ValueError(f"Usage line must be dict, got: {line}")

        ws.cell(row=r, column=unit_col,   value=safe_val(line.get("Units")))
        ws.cell(row=r, column=desc_col,   value=safe_val(line.get("Description")))
        ws.cell(row=r, column=before_col, value=safe_val(line.get("Rate")))
        ws.cell(row=r, column=disc_col,   value=safe_val(line.get("Discount")))

    # Save updated workbook to bytes
    output = io.BytesIO()
    wb.save(output)

    return output, "filled_quote.xlsx"


# -----------------------
# Streamlit UI
# -----------------------
st.set_page_config(page_title="Electricity Bill -> Quote Template", layout="wide")
st.title("⚡ Bill → Quote template (fixed Excel template)")

st.markdown(
    "Upload a bill (PDF/image). The app will extract header info and dynamic usage lines and fill them into your fixed Excel template.\n\n"
    "**Columns used:** Units → Column A; Description → Column C; Before Discount → Column D; Conditional Discount → Column E."
)

col1, col2 = st.columns([1, 2])

with col1:
    st.header("Step 1 — Upload files")
    bill_file = st.file_uploader("Upload bill (PDF, JPG, PNG)", type=["pdf", "jpg", "jpeg", "png"])
    st.markdown("**Template**: upload your fixed template (xlsx) or the repo's template will be used (if present).")
    template_file = st.file_uploader("Upload Excel template (.xlsx) (optional)", type=["xlsx"])

    st.markdown(
        "If you don't upload a template, the app expects a file named `Quote - (Site Address) - (Mth Year).xlsx` in the same folder as this app."
    )

with col2:
    st.header("Preview / Options")
    st.write("Tweak parser behaviour below if needed.")
    psm_choice = st.selectbox("Tesseract PSM (page segmentation)", options=["3","6","11"], index=1)
    st.caption("PSM helps OCR structure: 3 = fully automatic, 6 = assume a single uniform block, 11 = sparse text")
    # This would be passed into ocr_image cfg if we wired it (for now we use psm=6 in code)

process_btn = st.button("Extract & Parse")

if process_btn:
    if not bill_file:
        st.error("Please upload a bill file first.")
    else:
        file_bytes = bill_file.read()
        text = ""
        if bill_file.type == "application/pdf":
            text = extract_text_from_pdf_bytes(file_bytes)
        else:
            text = extract_text_from_image_bytes(file_bytes)

        st.subheader("Extracted text (preview)")
        st.text_area("Raw extracted text", value=text[:20000], height=300)

        # parse headers & usage lines
        headers = parse_header_fields(text)
        usage_lines = parse_usage_lines(text)

        st.subheader("Parsed header fields")
        st.json(headers)

        st.subheader("Parsed usage lines (editable)")
        if usage_lines:
            df = pd.DataFrame(usage_lines)
        else:
            # empty template for editing
            df = pd.DataFrame([{"Units": "", "Description": "", "Rate": "", "Discount": ""} for _ in range(5)])

        # allow user to edit the parsed table before writing
        try:
            edited = st.experimental_data_editor(df, num_rows="dynamic")
            # standardize column names -> convert to list of dicts
            usage_lines_final = edited.rename(columns={"Rate": "Rate", "Discount": "Discount", "Units": "Units", "Description": "Description"}).to_dict(orient="records")
        except Exception:
            # fallback to non-editable view
            st.dataframe(df)
            usage_lines_final = df.to_dict(orient="records")

        # template bytes
        if template_file:
            template_bytes = template_file.read()
        else:
            # attempt to read from a local filename
            try:
                with open("Quote - (Site Address) - (Mth Year).xlsx", "rb") as f:
                    template_bytes = f.read()
            except FileNotFoundError:
                st.error("No template uploaded and default template file not found in repo. Please upload your xlsx template.")
                template_bytes = None

        if template_bytes:
            filled_io, out_name = write_usage_to_template(template_bytes, headers, usage_lines_final)
            st.success("Template filled. Download below.")
            st.download_button("⬇️ Download filled template", data=filled_io, file_name=out_name, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        else:
            st.info("Upload a template to generate filled file.")
